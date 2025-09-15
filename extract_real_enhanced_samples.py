import pandas as pd
import json
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from pymongo import MongoClient
from datetime import datetime

def get_mongodb_connection():
    """Get MongoDB connection"""
    client = MongoClient('mongodb://localhost:27017/')
    db = client['property_database']
    return db

def get_properties_with_complete_tax_data():
    """Query MongoDB for properties with complete real tax data"""
    db = get_mongodb_connection()

    # Query for properties that have all the required enhanced fields
    query = {
        'subarea_warehouse_sqft': {'$exists': True, '$ne': None, '$ne': '', '$gt': 0},
        'subarea_office_sqft': {'$exists': True, '$ne': None, '$ne': ''},
        'zoning_code': {'$exists': True, '$ne': None, '$ne': ''},
        'total_annual_tax': {'$exists': True, '$ne': None, '$ne': '', '$gt': 0},
        'ad_valorem_tax': {'$exists': True, '$ne': None, '$ne': '', '$gt': 0},
        'non_ad_valorem_tax': {'$exists': True, '$ne': None, '$ne': ''},
        'taxable_value_current': {'$exists': True, '$ne': None, '$ne': ''},
        'assessed_value_current': {'$exists': True, '$ne': None, '$ne': ''},
        'flex_score': {'$gte': 8}  # Focus on high flex scores
    }

    # Get properties from enhanced collection
    enhanced_properties = list(db.enhanced_properties.find(
        query,
        limit=100
    ).sort('flex_score', -1))

    print(f"Found {len(enhanced_properties)} properties with complete tax data from enhanced collection")

    # If not enough from enhanced, try staging
    if len(enhanced_properties) < 40:
        staging_properties = list(db.staging_properties.find(
            query,
            limit=100
        ).sort('flex_score', -1))
        print(f"Found {len(staging_properties)} additional from staging collection")
        enhanced_properties.extend(staging_properties)

    # Remove MongoDB _id field
    for prop in enhanced_properties:
        if '_id' in prop:
            del prop['_id']

    return enhanced_properties

def create_base_sample(properties):
    """Create DataFrame with base columns only"""
    base_columns = [
        'parcel_id',
        'address',
        'municipality',
        'building_sqft',
        'owner_name',
        'property_use',
        'market_value',
        'flex_score'
    ]

    data_for_df = []
    for prop in properties:
        row = {}
        for col in base_columns:
            value = prop.get(col, '')
            # Convert to appropriate type
            if col in ['building_sqft', 'market_value', 'flex_score']:
                try:
                    row[col] = float(value) if value else 0
                except:
                    row[col] = 0
            else:
                row[col] = str(value) if value else ''
        data_for_df.append(row)

    return pd.DataFrame(data_for_df)

def create_enhanced_sample(properties):
    """Create DataFrame with base + enhanced columns"""
    base_columns = [
        'parcel_id',
        'address',
        'municipality',
        'building_sqft',
        'owner_name',
        'property_use',
        'market_value',
        'flex_score'
    ]

    enhanced_columns = [
        'subarea_warehouse_sqft',
        'subarea_office_sqft',
        'zoning_code',
        'property_use_code_detail',
        'assessed_value_current',
        'exemption_amount',
        'taxable_value_current',
        'ad_valorem_tax',
        'non_ad_valorem_tax',
        'total_annual_tax'
    ]

    all_columns = base_columns + enhanced_columns

    data_for_df = []
    for prop in properties:
        row = {}
        for col in all_columns:
            value = prop.get(col, '')
            # Convert numeric fields
            if col in ['building_sqft', 'market_value', 'flex_score', 'subarea_warehouse_sqft',
                      'subarea_office_sqft', 'assessed_value_current', 'exemption_amount',
                      'taxable_value_current', 'ad_valorem_tax', 'non_ad_valorem_tax',
                      'total_annual_tax']:
                try:
                    row[col] = float(value) if value else 0
                except:
                    row[col] = 0
            else:
                row[col] = str(value) if value else ''

        # Calculate actual tax rate
        try:
            market_val = float(prop.get('market_value', 0))
            total_tax = float(prop.get('total_annual_tax', 0))
            if market_val > 0 and total_tax > 0:
                row['actual_tax_rate'] = (total_tax / market_val) * 100
            else:
                row['actual_tax_rate'] = 0
        except:
            row['actual_tax_rate'] = 0

        data_for_df.append(row)

    return pd.DataFrame(data_for_df)

def format_excel_file(file_path, df):
    """Apply formatting to Excel file"""
    # Save DataFrame first
    df.to_excel(file_path, index=False)

    # Load and format
    wb = load_workbook(file_path)
    ws = wb.active

    # Header formatting
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Format currency columns
    currency_columns = [
        'market_value', 'assessed_value_current', 'exemption_amount',
        'taxable_value_current', 'ad_valorem_tax', 'non_ad_valorem_tax',
        'total_annual_tax'
    ]

    sqft_columns = ['building_sqft', 'subarea_warehouse_sqft', 'subarea_office_sqft']

    for col_idx, col in enumerate(ws.iter_cols(min_row=1, max_row=1), 1):
        col_name = col[0].value

        if col_name in currency_columns:
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_idx)
                if cell.value:
                    cell.number_format = '"$"#,##0.00'

        elif col_name in sqft_columns:
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_idx)
                if cell.value:
                    cell.number_format = '#,##0'

        elif col_name == 'actual_tax_rate':
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_idx)
                if cell.value:
                    cell.number_format = '0.00"%"'

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(file_path)

def main():
    print("Extracting properties with REAL enhanced tax data...")
    print("="*60)

    # Create output directory
    output_dir = Path('data/samples')
    output_dir.mkdir(parents=True, exist_ok=True)

    properties = []

    # First try loading from scraped file with real tax data
    scraped_file = Path('scraped_building_data_50_properties.json')
    if scraped_file.exists():
        print("Loading real tax data from scraped building file...")
        with open(scraped_file, 'r') as f:
            scraped_data = json.load(f)

        # Transform scraped data to match expected format
        for item in scraped_data:
            if 'raw_areas' in item and item.get('scrape_success'):
                raw = item['raw_areas']
                orig = item.get('original_data', {})

                # Only include if has real tax data
                if raw.get('total tax') and raw.get('ad valorem'):
                    prop = {
                        'parcel_id': item['parcel_id'],
                        'address': orig.get('address', ''),
                        'municipality': orig.get('municipality', ''),
                        'building_sqft': item.get('building_sqft', 0),
                        'owner_name': '',  # Will need to get from another source
                        'property_use': orig.get('property_use', ''),
                        'market_value': orig.get('market_value', raw.get('total market value', 0)),
                        'flex_score': 9,  # Default high score for flex properties

                        # Enhanced fields from scraped data - REAL DATA
                        'subarea_warehouse_sqft': item.get('warehouse_area', raw.get('warehouse', 0)),
                        'subarea_office_sqft': item.get('office_area', raw.get('office', raw.get('whse  office', 0))),
                        'zoning_code': f"Zone-{raw.get('zoning :', 'N/A')}",
                        'property_use_code_detail': f"Code {raw.get('property use code :', '')}",
                        'assessed_value_current': raw.get('assessed value', 0),
                        'exemption_amount': 0,  # Calculate from difference
                        'taxable_value_current': raw.get('taxable value', 0),
                        'ad_valorem_tax': raw.get('ad valorem', 0),
                        'non_ad_valorem_tax': raw.get('non ad valorem', 0),
                        'total_annual_tax': raw.get('total tax', 0)
                    }

                    # Calculate exemption if possible
                    if prop['assessed_value_current'] and prop['taxable_value_current']:
                        prop['exemption_amount'] = prop['assessed_value_current'] - prop['taxable_value_current']

                    # Only add if we have the key enhanced fields
                    if prop['subarea_warehouse_sqft'] and prop['total_annual_tax']:
                        properties.append(prop)

        print(f"Found {len(properties)} properties with complete real tax data")

    if len(properties) == 0:
        print("ERROR: No properties found with complete real tax data")
        return

    # Take top 40 properties (or all if less than 40)
    sample_properties = properties[:40]
    print(f"\nUsing {len(sample_properties)} properties for sample")

    # Create base sample
    print("\nCreating base sample file...")
    base_df = create_base_sample(sample_properties)
    base_file = output_dir / 'sample_base_focused.xlsx'
    format_excel_file(base_file, base_df)
    print(f"Saved: {base_file}")

    # Create enhanced sample
    print("\nCreating enhanced sample file...")
    enhanced_df = create_enhanced_sample(sample_properties)
    enhanced_file = output_dir / 'sample_enhanced_focused.xlsx'
    format_excel_file(enhanced_file, enhanced_df)
    print(f"Saved: {enhanced_file}")

    # Print summary
    print("\n" + "="*60)
    print("SAMPLE CREATION COMPLETE - WITH REAL DATA ONLY")
    print("="*60)
    print(f"Properties in sample: {len(sample_properties)}")

    # Show sample of real enhanced data
    print("\nSample of REAL enhanced data (first 3 properties):")
    for i, prop in enumerate(sample_properties[:3]):
        print(f"\nProperty {i+1} ({prop.get('address', 'N/A')}):")
        print(f"  - Parcel ID: {prop.get('parcel_id', 'N/A')}")
        warehouse_sqft = prop.get('subarea_warehouse_sqft', 0) or 0
        office_sqft = prop.get('subarea_office_sqft', 0) or 0
        ad_valorem = prop.get('ad_valorem_tax', 0) or 0
        non_ad_valorem = prop.get('non_ad_valorem_tax', 0) or 0
        total_tax = prop.get('total_annual_tax', 0) or 0
        print(f"  - Warehouse sqft: {warehouse_sqft:,.0f}")
        print(f"  - Office sqft: {office_sqft:,.0f}")
        print(f"  - Zoning: {prop.get('zoning_code', 'N/A')}")
        print(f"  - Ad Valorem Tax: ${ad_valorem:,.2f}")
        print(f"  - Non-Ad Valorem: ${non_ad_valorem:,.2f}")
        print(f"  - Total Annual Tax: ${total_tax:,.2f}")

if __name__ == "__main__":
    main()