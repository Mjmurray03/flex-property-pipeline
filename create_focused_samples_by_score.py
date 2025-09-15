import pandas as pd
import json
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

def load_complete_flex_properties():
    """Load the complete flex properties with scores"""
    complete_file = Path('data/exports/complete_flex_properties.json')
    with open(complete_file, 'r') as f:
        return json.load(f)

def load_scraped_tax_data():
    """Load real tax data from scraped building file"""
    scraped_file = Path('scraped_building_data_50_properties.json')
    tax_data = {}

    if scraped_file.exists():
        with open(scraped_file, 'r') as f:
            scraped_data = json.load(f)

        for item in scraped_data:
            if 'raw_areas' in item and item.get('scrape_success'):
                parcel_id = item['parcel_id']
                raw = item['raw_areas']

                # Only store if has real tax data
                if raw.get('total tax') and raw.get('ad valorem'):
                    tax_data[parcel_id] = {
                        'warehouse_area': item.get('warehouse_area', raw.get('warehouse', 0)),
                        'office_area': item.get('office_area', raw.get('office', raw.get('whse  office', 0))),
                        'zoning_code': f"Zone-{raw.get('zoning :', 'N/A')}",
                        'property_use_code_detail': f"Code {raw.get('property use code :', '')}",
                        'assessed_value_current': raw.get('assessed value', 0),
                        'taxable_value_current': raw.get('taxable value', 0),
                        'ad_valorem_tax': raw.get('ad valorem', 0),
                        'non_ad_valorem_tax': raw.get('non ad valorem', 0),
                        'total_annual_tax': raw.get('total tax', 0),
                        'improvement_value': raw.get('improvement value', 0),
                        'land_value': raw.get('land value', 0),
                        'total_market_value': raw.get('total market value', 0)
                    }

    return tax_data

def select_properties_by_score(properties, tax_data):
    """Select 20 properties with score 10 and 20 with score 9"""
    score_10_props = [p for p in properties if p.get('flex_score') == 10]
    score_9_props = [p for p in properties if p.get('flex_score') == 9]

    print(f"Available properties with score 10: {len(score_10_props)}")
    print(f"Available properties with score 9: {len(score_9_props)}")

    # Prioritize properties that have tax data
    score_10_with_tax = [p for p in score_10_props if p.get('parcel_id') in tax_data]
    score_10_without_tax = [p for p in score_10_props if p.get('parcel_id') not in tax_data]

    score_9_with_tax = [p for p in score_9_props if p.get('parcel_id') in tax_data]
    score_9_without_tax = [p for p in score_9_props if p.get('parcel_id') not in tax_data]

    print(f"Score 10 properties with tax data: {len(score_10_with_tax)}")
    print(f"Score 9 properties with tax data: {len(score_9_with_tax)}")

    # Select 20 of each, prioritizing those with tax data
    selected_10s = (score_10_with_tax + score_10_without_tax)[:20]
    selected_9s = (score_9_with_tax + score_9_without_tax)[:20]

    return selected_10s + selected_9s

def enhance_property_with_real_data(prop, tax_data):
    """Enhance property with real tax data if available, use analytical estimates otherwise"""
    from analytical_tax_estimator import PalmBeachCountyTaxEstimator

    enhanced_prop = prop.copy()
    parcel_id = prop.get('parcel_id')

    # Always start with existing warehouse/office data from base property
    enhanced_prop['subarea_warehouse_sqft'] = prop.get('warehouse_sqft', 0) or 0
    enhanced_prop['subarea_office_sqft'] = prop.get('office_sqft', 0) or 0

    if parcel_id in tax_data:
        # REAL DATA - Use actual scraped tax records
        real_data = tax_data[parcel_id]

        # Override with real scraped data if available and non-zero
        if real_data['warehouse_area']:
            enhanced_prop['subarea_warehouse_sqft'] = real_data['warehouse_area']
        if real_data['office_area']:
            enhanced_prop['subarea_office_sqft'] = real_data['office_area']

        # Add real enhanced fields
        enhanced_prop['zoning_code'] = real_data['zoning_code']
        enhanced_prop['property_use_code_detail'] = real_data['property_use_code_detail']
        enhanced_prop['assessed_value_current'] = real_data['assessed_value_current']
        enhanced_prop['taxable_value_current'] = real_data['taxable_value_current']
        enhanced_prop['ad_valorem_tax'] = real_data['ad_valorem_tax']
        enhanced_prop['non_ad_valorem_tax'] = real_data['non_ad_valorem_tax']
        enhanced_prop['total_annual_tax'] = real_data['total_annual_tax']

        # Calculate exemption from real data
        assessed = real_data['assessed_value_current']
        taxable = real_data['taxable_value_current']
        enhanced_prop['exemption_amount'] = max(0, assessed - taxable) if assessed and taxable else 0

        # Use real market value if available and different from base
        if real_data['total_market_value'] and real_data['total_market_value'] != prop.get('market_value'):
            enhanced_prop['market_value'] = real_data['total_market_value']

        # Mark as real data
        enhanced_prop['data_source'] = 'REAL_SCRAPED_DATA'

    else:
        # ANALYTICAL ESTIMATES - Use first-principles calculations based on empirical analysis
        market_value = prop.get('market_value', 0) or 0
        property_use = prop.get('property_use', '')
        municipality = prop.get('municipality', '')
        building_sqft = prop.get('building_sqft', 0) or 0

        if market_value > 0:
            # Generate complete analytical tax profile
            tax_profile = PalmBeachCountyTaxEstimator.estimate_complete_tax_profile(
                market_value=market_value,
                property_use=property_use,
                municipality=municipality,
                building_sqft=building_sqft
            )

            # Extract calculated values
            calc_values = tax_profile['calculated_values']
            enhanced_prop['assessed_value_current'] = calc_values['assessed_value_current']
            enhanced_prop['exemption_amount'] = calc_values['exemption_amount']
            enhanced_prop['taxable_value_current'] = calc_values['taxable_value_current']
            enhanced_prop['ad_valorem_tax'] = calc_values['ad_valorem_tax']
            enhanced_prop['non_ad_valorem_tax'] = calc_values['non_ad_valorem_tax']
            enhanced_prop['total_annual_tax'] = calc_values['total_annual_tax']

            # Add methodology documentation
            enhanced_prop['data_source'] = 'ANALYTICAL_ESTIMATE'
            enhanced_prop['estimation_methodology'] = tax_profile['methodology']

        else:
            # Fallback for properties with no market value
            enhanced_prop['assessed_value_current'] = 0
            enhanced_prop['exemption_amount'] = 0
            enhanced_prop['taxable_value_current'] = 0
            enhanced_prop['ad_valorem_tax'] = 0
            enhanced_prop['non_ad_valorem_tax'] = 0
            enhanced_prop['total_annual_tax'] = 0
            enhanced_prop['data_source'] = 'INSUFFICIENT_DATA'

        # Estimate zoning and property details
        enhanced_prop['zoning_code'] = get_estimated_zoning(property_use)
        enhanced_prop['property_use_code_detail'] = property_use

    return enhanced_prop

def get_estimated_zoning(property_use):
    """Get estimated zoning based on property use"""
    if not property_use:
        return "Unknown"

    property_use_upper = property_use.upper()
    if 'WAREH' in property_use_upper or 'DIST' in property_use_upper:
        return "IL"  # Industrial Light
    elif 'OFFICE' in property_use_upper:
        return "PID"  # Planned Industrial Development
    elif 'FLEX' in property_use_upper or 'TECH' in property_use_upper:
        return "IG"  # Industrial General
    elif 'MANUF' in property_use_upper:
        return "IH"  # Industrial Heavy
    else:
        return "IL"  # Default to Industrial Light

def create_base_sample(properties):
    """Create DataFrame with base columns only"""
    base_columns = [
        'parcel_id',
        'address',
        'municipality',
        'building_sqft',
        'owner_name',
        'property_use',
        'market_value',
        'flex_score'
    ]

    data_for_df = []
    for prop in properties:
        row = {}
        for col in base_columns:
            value = prop.get(col, '')
            # Convert to appropriate type
            if col in ['building_sqft', 'market_value', 'flex_score']:
                try:
                    row[col] = float(value) if value else 0
                except:
                    row[col] = 0
            else:
                row[col] = str(value) if value else ''
        data_for_df.append(row)

    return pd.DataFrame(data_for_df)

def create_enhanced_sample(properties):
    """Create DataFrame with base + enhanced columns"""
    base_columns = [
        'parcel_id',
        'address',
        'municipality',
        'building_sqft',
        'owner_name',
        'property_use',
        'market_value',
        'flex_score'
    ]

    enhanced_columns = [
        'subarea_warehouse_sqft',
        'subarea_office_sqft',
        'zoning_code',
        'property_use_code_detail',
        'assessed_value_current',
        'exemption_amount',
        'taxable_value_current',
        'ad_valorem_tax',
        'non_ad_valorem_tax',
        'total_annual_tax',
        'data_source'
    ]

    all_columns = base_columns + enhanced_columns

    data_for_df = []
    for prop in properties:
        row = {}
        for col in all_columns:
            value = prop.get(col, '')

            # Handle numeric fields, but preserve "Data Not Available" strings
            if col in ['building_sqft', 'market_value', 'flex_score', 'subarea_warehouse_sqft',
                      'subarea_office_sqft', 'assessed_value_current', 'exemption_amount',
                      'taxable_value_current']:
                try:
                    if value == "Data Not Available" or value == '':
                        row[col] = value
                    else:
                        row[col] = float(value) if value else 0
                except:
                    row[col] = value
            elif col in ['ad_valorem_tax', 'non_ad_valorem_tax', 'total_annual_tax']:
                # These might be "Data Not Available" strings or numbers
                try:
                    if value == "Data Not Available" or value == '':
                        row[col] = value
                    else:
                        row[col] = float(value) if value else 0
                except:
                    row[col] = value
            else:
                row[col] = str(value) if value else ''

        # Calculate actual tax rate only if we have numeric data
        try:
            market_val = float(prop.get('market_value', 0))
            total_tax = prop.get('total_annual_tax', 0)
            if isinstance(total_tax, (int, float)) and market_val > 0 and total_tax > 0:
                row['actual_tax_rate'] = (total_tax / market_val) * 100
            else:
                row['actual_tax_rate'] = "Data Not Available"
        except:
            row['actual_tax_rate'] = "Data Not Available"

        data_for_df.append(row)

    return pd.DataFrame(data_for_df)

def format_excel_file(file_path, df):
    """Apply formatting to Excel file"""
    # Save DataFrame first
    df.to_excel(file_path, index=False)

    # Load and format
    wb = load_workbook(file_path)
    ws = wb.active

    # Header formatting
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Format currency columns
    currency_columns = [
        'market_value', 'assessed_value_current', 'exemption_amount',
        'taxable_value_current', 'ad_valorem_tax', 'non_ad_valorem_tax',
        'total_annual_tax'
    ]

    sqft_columns = ['building_sqft', 'subarea_warehouse_sqft', 'subarea_office_sqft']

    for col_idx, col in enumerate(ws.iter_cols(min_row=1, max_row=1), 1):
        col_name = col[0].value

        if col_name in currency_columns:
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_idx)
                if cell.value and isinstance(cell.value, (int, float)):
                    cell.number_format = '"$"#,##0.00'

        elif col_name in sqft_columns:
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_idx)
                if cell.value and isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0'

        elif col_name == 'actual_tax_rate':
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_idx)
                if cell.value and isinstance(cell.value, (int, float)):
                    cell.number_format = '0.00"%"'

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(file_path)

def main():
    print("Creating focused samples: 20 properties with score 10 + 20 with score 9")
    print("="*70)

    # Create output directory
    output_dir = Path('data/samples')
    output_dir.mkdir(parents=True, exist_ok=True)

    # Load complete flex properties dataset
    print("Loading complete flex properties dataset...")
    all_properties = load_complete_flex_properties()
    print(f"Loaded {len(all_properties)} total properties")

    # Load real tax data
    print("Loading real tax data from scraped building file...")
    tax_data = load_scraped_tax_data()
    print(f"Found real tax data for {len(tax_data)} properties")

    # Select properties by score
    print("\nSelecting properties by flex score...")
    selected_properties = select_properties_by_score(all_properties, tax_data)
    print(f"Selected {len(selected_properties)} properties total")

    # Enhance with real data
    print("Enhancing properties with real tax data...")
    enhanced_properties = []
    properties_with_real_tax = 0

    for prop in selected_properties:
        enhanced_prop = enhance_property_with_real_data(prop, tax_data)
        enhanced_properties.append(enhanced_prop)

        if prop.get('parcel_id') in tax_data:
            properties_with_real_tax += 1

    print(f"Properties with complete real tax data: {properties_with_real_tax}")
    print(f"Properties with partial data: {len(enhanced_properties) - properties_with_real_tax}")

    # Create base sample
    print("\nCreating base sample file...")
    base_df = create_base_sample(enhanced_properties)
    base_file = output_dir / 'sample_base_focused.xlsx'
    format_excel_file(base_file, base_df)
    print(f"Saved: {base_file}")

    # Create enhanced sample
    print("\nCreating enhanced sample file...")
    enhanced_df = create_enhanced_sample(enhanced_properties)
    enhanced_file = output_dir / 'sample_enhanced_focused.xlsx'
    format_excel_file(enhanced_file, enhanced_df)
    print(f"Saved: {enhanced_file}")

    # Print summary
    print("\n" + "="*70)
    print("SAMPLE CREATION COMPLETE - 20 TENS + 20 NINES")
    print("="*70)
    print(f"Total properties: {len(enhanced_properties)}")

    score_10_count = len([p for p in enhanced_properties if p.get('flex_score') == 10])
    score_9_count = len([p for p in enhanced_properties if p.get('flex_score') == 9])
    print(f"Score 10 properties: {score_10_count}")
    print(f"Score 9 properties: {score_9_count}")
    print(f"Properties with real tax data: {properties_with_real_tax}")

    # Show sample of enhanced data
    print("\nSample of enhanced data (first 3 properties):")
    for i, prop in enumerate(enhanced_properties[:3]):
        print(f"\nProperty {i+1} ({prop.get('address', 'N/A')}):")
        print(f"  - Parcel ID: {prop.get('parcel_id', 'N/A')}")
        print(f"  - Flex Score: {prop.get('flex_score', 'N/A')}")

        warehouse_sqft = prop.get('subarea_warehouse_sqft', 0)
        office_sqft = prop.get('subarea_office_sqft', 0)
        print(f"  - Warehouse sqft: {warehouse_sqft:,.0f}" if isinstance(warehouse_sqft, (int, float)) else f"  - Warehouse sqft: {warehouse_sqft}")
        print(f"  - Office sqft: {office_sqft:,.0f}" if isinstance(office_sqft, (int, float)) else f"  - Office sqft: {office_sqft}")
        print(f"  - Zoning: {prop.get('zoning_code', 'N/A')}")

        total_tax = prop.get('total_annual_tax', 0)
        if isinstance(total_tax, (int, float)):
            print(f"  - Total Annual Tax: ${total_tax:,.2f}")
        else:
            print(f"  - Total Annual Tax: {total_tax}")

if __name__ == "__main__":
    main()