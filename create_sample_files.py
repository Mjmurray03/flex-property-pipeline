#!/usr/bin/env python3
"""
Create Sample Excel Files - Base vs Enhanced Property Data Comparison
"""

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, NamedStyle
from openpyxl.formatting import Rule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.utils.dataframe import dataframe_to_rows
import os

def create_sample_files():
    """Create professional sample Excel files showing base vs enhanced data"""

    print("=== CREATING SAMPLE EXCEL FILES ===")

    # Create samples directory
    sample_dir = 'data/samples'
    os.makedirs(sample_dir, exist_ok=True)

    # Load enhanced dataset
    enhanced_df = pd.read_excel('data/exports/complete_flex_properties_ENHANCED.xlsx')
    print(f"Loaded {len(enhanced_df)} enhanced properties")

    # Extract 40 sample properties (20 with score 10, 20 with score 9)
    score_10_properties = enhanced_df[enhanced_df['flex_score'] == 10].head(20)
    score_9_properties = enhanced_df[enhanced_df['flex_score'] == 9].head(20)

    sample_properties = pd.concat([score_10_properties, score_9_properties])
    print(f"Selected {len(sample_properties)} sample properties (20 score 10, 20 score 9)")

    # =================================================================
    # FILE 1: BASE PROPERTIES SAMPLE
    # =================================================================

    print("\nCreating sample_base_properties.xlsx...")

    # Select base columns only
    base_columns = [
        'parcel_id', 'address', 'municipality',
        'building_sqft', 'year_built',
        'owner_name', 'owner_address',
        'property_use', 'acres',
        'market_value', 'assessed_value',
        'sale_date', 'sale_price',
        'flex_score'
    ]

    base_sample = sample_properties[base_columns].copy()

    # Create base Excel file with formatting
    create_formatted_excel(
        base_sample,
        os.path.join(sample_dir, 'sample_base_properties.xlsx'),
        "Base Property Data Sample",
        base_columns=True
    )

    # =================================================================
    # FILE 2: ENHANCED PROPERTIES SAMPLE
    # =================================================================

    print("Creating sample_enhanced_properties.xlsx...")

    # Calculate additional fields for enhanced sample
    enhanced_sample = sample_properties.copy()

    # Use existing 5yr_appreciation if available, otherwise set to None
    if '5yr_appreciation' in enhanced_sample.columns:
        enhanced_sample['5yr_appreciation_percent'] = enhanced_sample['5yr_appreciation']
    else:
        enhanced_sample['5yr_appreciation_percent'] = None

    # Use existing tax_rate if available, otherwise calculate from current data
    if 'tax_rate' in enhanced_sample.columns:
        enhanced_sample['tax_rate_percent'] = enhanced_sample['tax_rate']
    else:
        enhanced_sample['tax_rate_percent'] = None
        valid_tax_rate = (
            enhanced_sample['total_annual_tax'].notna() &
            enhanced_sample['market_value'].notna() &
            (enhanced_sample['market_value'] > 0)
        )
        enhanced_sample.loc[valid_tax_rate, 'tax_rate_percent'] = (
            enhanced_sample.loc[valid_tax_rate, 'total_annual_tax'] /
            enhanced_sample.loc[valid_tax_rate, 'market_value'] * 100
        ).round(3)

    # Calculate number_of_sales
    enhanced_sample['number_of_sales'] = enhanced_sample['sales_history'].apply(
        lambda x: len(x) if isinstance(x, list) else 0
    )

    # Calculate last_sale_days_ago
    enhanced_sample['last_sale_days_ago'] = None

    def calculate_days_since_sale(sale_date_str):
        if pd.isna(sale_date_str) or sale_date_str == '':
            return None
        try:
            # Try multiple date formats
            for fmt in ['%m/%d/%Y', '%Y-%m-%d', '%m-%d-%Y']:
                try:
                    sale_date = datetime.strptime(str(sale_date_str), fmt)
                    days_ago = (datetime.now() - sale_date).days
                    return days_ago
                except ValueError:
                    continue
            return None
        except:
            return None

    enhanced_sample['last_sale_days_ago'] = enhanced_sample['sale_date'].apply(calculate_days_since_sale)

    # Select enhanced columns
    enhanced_columns = [
        # Base identification
        'parcel_id', 'address', 'municipality',
        'building_sqft', 'year_built',
        'owner_name', 'owner_address',
        'property_use', 'acres',
        'market_value', 'assessed_value',
        'sale_date', 'sale_price',
        'flex_score',

        # Enhanced property details
        'subarea_warehouse_sqft', 'subarea_office_sqft',
        'office_warehouse_ratio',
        'zoning_code',

        # Enhanced valuations
        'improvement_value', 'land_value',
        '5yr_appreciation_percent',

        # Enhanced tax data
        'total_annual_tax', 'tax_rate_percent',

        # Enhanced sales data
        'number_of_sales', 'last_sale_days_ago'
    ]

    enhanced_sample_final = enhanced_sample[enhanced_columns].copy()

    # Create enhanced Excel file with formatting
    create_formatted_excel(
        enhanced_sample_final,
        os.path.join(sample_dir, 'sample_enhanced_properties.xlsx'),
        "Enhanced Property Data Sample",
        base_columns=False
    )

    print(f"\nSample files created successfully:")
    print(f"   {sample_dir}/sample_base_properties.xlsx")
    print(f"   {sample_dir}/sample_enhanced_properties.xlsx")

    # Print summary
    print(f"\nSAMPLE SUMMARY:")
    print(f"   Total properties: {len(sample_properties)}")
    print(f"   Properties with score 10: {len(score_10_properties)}")
    print(f"   Properties with score 9: {len(score_9_properties)}")
    print(f"   Base columns: {len(base_columns)}")
    print(f"   Enhanced columns: {len(enhanced_columns)}")
    print(f"   Additional fields: {len(enhanced_columns) - len(base_columns)}")

def create_formatted_excel(df, filename, sheet_name, base_columns=True):
    """Create professionally formatted Excel file"""

    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Add dataframe to worksheet
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # =================================================================
    # FORMATTING
    # =================================================================

    # Header formatting
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    # Freeze top row
    ws.freeze_panes = "A2"

    # Column formatting based on data type
    currency_columns = ['market_value', 'assessed_value', 'sale_price', 'total_annual_tax',
                       'improvement_value', 'land_value']
    percentage_columns = ['office_warehouse_ratio', '5yr_appreciation_percent', 'tax_rate_percent']
    number_columns = ['building_sqft', 'subarea_warehouse_sqft', 'subarea_office_sqft', 'year_built',
                     'number_of_sales', 'last_sale_days_ago']

    # Apply formatting to data columns
    for col_idx, column in enumerate(df.columns, 1):
        col_letter = chr(64 + col_idx)  # A, B, C, etc.

        # Currency formatting
        if column in currency_columns:
            for row in range(2, len(df) + 2):
                cell = ws[f"{col_letter}{row}"]
                if cell.value is not None and cell.value != '':
                    try:
                        cell.number_format = '"$"#,##0'
                    except:
                        pass

        # Percentage formatting
        elif column in percentage_columns:
            for row in range(2, len(df) + 2):
                cell = ws[f"{col_letter}{row}"]
                if cell.value is not None and cell.value != '':
                    try:
                        cell.number_format = '0.00%'
                        # Convert decimal to percentage for display
                        if isinstance(cell.value, (int, float)) and cell.value > 1:
                            cell.value = cell.value / 100
                    except:
                        pass

        # Number formatting
        elif column in number_columns:
            for row in range(2, len(df) + 2):
                cell = ws[f"{col_letter}{row}"]
                if cell.value is not None and cell.value != '':
                    try:
                        cell.number_format = '#,##0'
                    except:
                        pass

    # =================================================================
    # CONDITIONAL FORMATTING - Flex Score
    # =================================================================

    # Find flex_score column
    flex_score_col = None
    for col_idx, column in enumerate(df.columns, 1):
        if column == 'flex_score':
            flex_score_col = chr(64 + col_idx)
            break

    if flex_score_col:
        # Green fill for score 10
        green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        green_rule = CellIsRule(operator='equal', formula=['10'], fill=green_fill)

        # Yellow fill for score 9
        yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        yellow_rule = CellIsRule(operator='equal', formula=['9'], fill=yellow_fill)

        # Apply conditional formatting to flex_score column
        score_range = f"{flex_score_col}2:{flex_score_col}{len(df) + 1}"
        ws.conditional_formatting.add(score_range, green_rule)
        ws.conditional_formatting.add(score_range, yellow_rule)

    # =================================================================
    # COLUMN WIDTH ADJUSTMENT
    # =================================================================

    column_widths = {
        'parcel_id': 18,
        'address': 35,
        'municipality': 15,
        'building_sqft': 12,
        'year_built': 10,
        'owner_name': 25,
        'owner_address': 30,
        'property_use': 18,
        'acres': 8,
        'market_value': 12,
        'assessed_value': 12,
        'sale_date': 12,
        'sale_price': 12,
        'flex_score': 10,
        'subarea_warehouse_sqft': 15,
        'subarea_office_sqft': 15,
        'office_warehouse_ratio': 15,
        'zoning_code': 20,
        'improvement_value': 15,
        'land_value': 15,
        '5yr_appreciation_percent': 15,
        'total_annual_tax': 15,
        'tax_rate_percent': 12,
        'number_of_sales': 12,
        'last_sale_days_ago': 15
    }

    for col_idx, column in enumerate(df.columns, 1):
        col_letter = chr(64 + col_idx)
        width = column_widths.get(column, 12)
        ws.column_dimensions[col_letter].width = width

    # Save workbook
    wb.save(filename)
    print(f"   Created: {filename}")

if __name__ == "__main__":
    create_sample_files()