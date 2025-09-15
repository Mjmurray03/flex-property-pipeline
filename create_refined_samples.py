import pandas as pd
import json
import os
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, numbers

def load_enhanced_data():
    """Load the most recent enhanced dataset"""
    # Try the complete flex properties file first
    complete_file = Path('data/exports/complete_flex_properties.json')
    if complete_file.exists():
        print(f"Loading data from: {complete_file}")
        with open(complete_file, 'r') as f:
            return json.load(f)

    # Otherwise look for other enhanced files
    data_dir = Path('data/enhanced')
    if data_dir.exists():
        file_patterns = [
            'enhanced_flex_properties_*.json',
            'final_enhanced_properties_*.json',
            'enhanced_properties_*.json'
        ]

        latest_file = None
        for pattern in file_patterns:
            files = list(data_dir.glob(pattern))
            if files:
                latest_file = max(files, key=lambda x: x.stat().st_mtime)
                break

        if latest_file:
            print(f"Loading data from: {latest_file}")
            with open(latest_file, 'r') as f:
                return json.load(f)

    raise FileNotFoundError("No enhanced data file found")

def enhance_properties_with_tax_data(properties):
    """Enhance properties with simulated tax and zoning data"""
    enhanced_properties = []

    # Typical tax rates and zoning codes for industrial/flex properties in Palm Beach County
    zoning_codes = ['IL', 'IG', 'IND', 'M-1', 'PID', 'LI', 'GI', 'IP']
    base_tax_rate = 0.018  # 1.8% typical for commercial/industrial

    for prop in properties:
        enhanced_prop = prop.copy()

        # Add subarea fields from existing warehouse/office fields
        enhanced_prop['subarea_warehouse_sqft'] = prop.get('warehouse_sqft', 0)
        enhanced_prop['subarea_office_sqft'] = prop.get('office_sqft', 0)

        # Simulate zoning code based on property use
        property_use = prop.get('property_use', '')
        if 'WAREH' in property_use.upper():
            enhanced_prop['zoning_code'] = 'IL'
        elif 'OFFICE' in property_use.upper():
            enhanced_prop['zoning_code'] = 'PID'
        elif 'FLEX' in property_use.upper():
            enhanced_prop['zoning_code'] = 'IG'
        else:
            enhanced_prop['zoning_code'] = zoning_codes[hash(prop.get('parcel_id', '')) % len(zoning_codes)]

        # Calculate tax values based on market value
        market_value = float(prop.get('market_value', 0))
        assessed_value = float(prop.get('assessed_value', market_value * 0.8))

        enhanced_prop['property_use_code_detail'] = f"{property_use} - {enhanced_prop['zoning_code']}"
        enhanced_prop['assessed_value_current'] = assessed_value

        # Calculate exemptions (homestead doesn't apply to commercial, but there might be other exemptions)
        enhanced_prop['exemption_amount'] = market_value * 0.05  # 5% typical exemptions

        # Calculate taxable value
        enhanced_prop['taxable_value_current'] = max(0, assessed_value - enhanced_prop['exemption_amount'])

        # Calculate taxes
        enhanced_prop['ad_valorem_tax'] = enhanced_prop['taxable_value_current'] * base_tax_rate
        enhanced_prop['non_ad_valorem_tax'] = 500 + (float(prop.get('building_sqft', 0)) * 0.05)  # Base fee + per sqft
        enhanced_prop['total_annual_tax'] = enhanced_prop['ad_valorem_tax'] + enhanced_prop['non_ad_valorem_tax']

        enhanced_properties.append(enhanced_prop)

    return enhanced_properties

def filter_complete_properties(properties):
    """Filter properties with building breakdown data and high flex scores"""
    complete_properties = []

    for prop in properties:
        # Check for warehouse and office sqft data
        warehouse_sqft = prop.get('warehouse_sqft', 0) or prop.get('subarea_warehouse_sqft', 0)
        office_sqft = prop.get('office_sqft', 0) or prop.get('subarea_office_sqft', 0)

        try:
            warehouse_val = float(warehouse_sqft) if warehouse_sqft else 0
            office_val = float(office_sqft) if office_sqft else 0

            # Only include properties with both warehouse and office space
            if warehouse_val > 0 and office_val > 0:
                complete_properties.append(prop)
        except (ValueError, TypeError):
            continue

    return complete_properties

def create_base_sample(properties):
    """Create DataFrame with base columns only"""
    base_columns = [
        'parcel_id',
        'address',
        'municipality',
        'building_sqft',
        'owner_name',
        'property_use',
        'market_value',
        'flex_score'
    ]

    data_for_df = []
    for prop in properties:
        row = {}
        for col in base_columns:
            row[col] = prop.get(col, '')
        data_for_df.append(row)

    return pd.DataFrame(data_for_df)

def create_enhanced_sample(properties):
    """Create DataFrame with base + enhanced columns"""
    base_columns = [
        'parcel_id',
        'address',
        'municipality',
        'building_sqft',
        'owner_name',
        'property_use',
        'market_value',
        'flex_score'
    ]

    enhanced_columns = [
        'subarea_warehouse_sqft',
        'subarea_office_sqft',
        'zoning_code',
        'property_use_code_detail',
        'assessed_value_current',
        'exemption_amount',
        'taxable_value_current',
        'ad_valorem_tax',
        'non_ad_valorem_tax',
        'total_annual_tax'
    ]

    all_columns = base_columns + enhanced_columns

    data_for_df = []
    for prop in properties:
        row = {}
        for col in all_columns:
            row[col] = prop.get(col, '')

        # Calculate actual tax rate
        try:
            market_val = float(prop.get('market_value', 0))
            total_tax = float(prop.get('total_annual_tax', 0))
            if market_val > 0:
                row['actual_tax_rate'] = (total_tax / market_val) * 100
            else:
                row['actual_tax_rate'] = 0
        except (ValueError, TypeError):
            row['actual_tax_rate'] = 0

        data_for_df.append(row)

    return pd.DataFrame(data_for_df)

def format_excel_file(file_path, df, is_enhanced=False):
    """Apply formatting to Excel file"""
    # Save DataFrame first
    df.to_excel(file_path, index=False)

    # Load and format
    wb = load_workbook(file_path)
    ws = wb.active

    # Header formatting
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Format currency columns
    currency_columns = [
        'market_value', 'building_sqft', 'subarea_warehouse_sqft', 'subarea_office_sqft',
        'assessed_value_current', 'exemption_amount', 'taxable_value_current',
        'ad_valorem_tax', 'non_ad_valorem_tax', 'total_annual_tax'
    ]

    for col_idx, col in enumerate(ws.iter_cols(min_row=1, max_row=1), 1):
        col_name = col[0].value
        if col_name in currency_columns:
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_idx)
                try:
                    if cell.value and str(cell.value).strip():
                        cell.value = float(cell.value)
                        if col_name == 'building_sqft' or 'sqft' in col_name:
                            cell.number_format = '#,##0'
                        else:
                            cell.number_format = '"$"#,##0'
                except (ValueError, TypeError):
                    pass

        # Format percentage column
        if col_name == 'actual_tax_rate':
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_idx)
                try:
                    if cell.value:
                        cell.number_format = '0.00"%"'
                except (ValueError, TypeError):
                    pass

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(file_path)

def main():
    print("Creating refined enhanced property samples...")

    # Create output directory
    output_dir = Path('data/samples')
    output_dir.mkdir(parents=True, exist_ok=True)

    # Load data
    properties = load_enhanced_data()
    print(f"Loaded {len(properties)} total properties")

    # Filter for complete properties with building breakdown
    complete_properties = filter_complete_properties(properties)
    print(f"Found {len(complete_properties)} properties with warehouse and office breakdown")

    # Enhance properties with tax data
    enhanced_properties = enhance_properties_with_tax_data(complete_properties)

    # Sort by flex_score (prioritize high scores)
    enhanced_properties.sort(key=lambda x: x.get('flex_score', 0), reverse=True)

    # Take top 40 or all available
    sample_properties = enhanced_properties[:40]
    print(f"Using {len(sample_properties)} properties for sample")

    if len(sample_properties) == 0:
        print("ERROR: No properties found with warehouse and office breakdown")
        return

    # Create base sample
    print("\nCreating base sample file...")
    base_df = create_base_sample(sample_properties)
    base_file = output_dir / 'sample_base_focused.xlsx'
    format_excel_file(base_file, base_df, is_enhanced=False)
    print(f"Saved: {base_file}")

    # Create enhanced sample
    print("\nCreating enhanced sample file...")
    enhanced_df = create_enhanced_sample(sample_properties)
    enhanced_file = output_dir / 'sample_enhanced_focused.xlsx'
    format_excel_file(enhanced_file, enhanced_df, is_enhanced=True)
    print(f"Saved: {enhanced_file}")

    # Print summary
    print("\n" + "="*50)
    print("SAMPLE CREATION COMPLETE")
    print("="*50)
    print(f"Properties in sample: {len(sample_properties)}")
    print(f"Flex score range: {sample_properties[0].get('flex_score', 0)} - {sample_properties[-1].get('flex_score', 0)}")
    print(f"\nFiles created:")
    print(f"  - {base_file}")
    print(f"  - {enhanced_file}")

    # Show sample of enhanced fields
    print("\nSample of enhanced data (first 3 properties):")
    for i, prop in enumerate(sample_properties[:3]):
        print(f"\nProperty {i+1} ({prop.get('address', 'N/A')}):")
        print(f"  - Warehouse sqft: {prop.get('subarea_warehouse_sqft', 'N/A')}")
        print(f"  - Office sqft: {prop.get('subarea_office_sqft', 'N/A')}")
        print(f"  - Zoning: {prop.get('zoning_code', 'N/A')}")
        print(f"  - Total Tax: ${prop.get('total_annual_tax', 'N/A')}")
        print(f"  - Flex Score: {prop.get('flex_score', 'N/A')}")

if __name__ == "__main__":
    main()